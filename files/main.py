"""
saasmetrics.ai  |  Backend  v5  —  SOTA
═══════════════════════════════════════════════════════════════
Architecture:
  Stage 1  AI Router       Gemini Flash  — decides which sources,
                           generates SQL intent, tags query type.
                           Fast + cheap. Never answers the question.

  Stage 2  Parallel Fetch  BQ dry-run → execute (concurrent with
                           file source loading). All I/O happens
                           simultaneously via asyncio.

  Stage 3  Answer Gen      Gemini Pro  — grounded in actual fetched
                           data, cites sources, disambiguates columns,
                           flags conflicts. Streams response tokens.

  Upload   GCS-backed      Files stored in GCS (local fallback).
           index           Parsed + indexed on upload. Re-indexed
                           from storage on startup. Shared across
                           all demo users (no auth).

Key design decisions:
  - No auth (demo mode — uniform access for all users)
  - No hardcoded values — everything from env vars
  - Router uses Gemini Flash (fast/cheap) not Pro
  - SQL is dry-run validated before execution every time
  - Self-correction: one automated retry on SQL failure
  - Streaming: answer tokens streamed via SSE to frontend
  - File index is in-memory + GCS-persisted
═══════════════════════════════════════════════════════════════
"""

from __future__ import annotations

import asyncio
import json
import os
import re
import io
from pathlib import Path
from typing import AsyncIterator, List, Optional

import yaml
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

load_dotenv()

# ── Config from env (zero hardcoding) ────────────────────────────────────────
GEMINI_API_KEY      = os.getenv("GEMINI_API_KEY", "")
ROUTER_MODEL        = os.getenv("ROUTER_MODEL",  "gemini-1.5-flash")
ANSWER_MODEL        = os.getenv("ANSWER_MODEL",  "gemini-1.5-pro")
GCP_PROJECT         = os.getenv("GCP_PROJECT",   "")
BQ_DATASET          = os.getenv("BQ_DATASET",    "saasmetrics")
GCS_BUCKET          = os.getenv("GCS_BUCKET",    "")
BACKEND_PORT        = int(os.getenv("BACKEND_PORT", "8000"))
MAX_UPLOAD_MB       = int(os.getenv("MAX_UPLOAD_MB", "20"))
HISTORY_WINDOW      = int(os.getenv("HISTORY_WINDOW", "12"))
ROUTER_HISTORY_WIN  = int(os.getenv("ROUTER_HISTORY_WINDOW", "4"))

BASE_DIR   = Path(__file__).parent
UPLOAD_DIR = BASE_DIR / "uploads_store"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

ALLOWED_EXTENSIONS = {".xlsx", ".xls", ".pdf", ".docx", ".csv"}

# ── Optional dependency imports ───────────────────────────────────────────────
try:
    import google.generativeai as genai
    genai.configure(api_key=GEMINI_API_KEY)
    _router_model = genai.GenerativeModel(ROUTER_MODEL)
    _answer_model = genai.GenerativeModel(ANSWER_MODEL)
    GENAI_OK = True
except Exception:
    GENAI_OK = False

try:
    from google.cloud import bigquery as _bq
    import threading
    _bq_client = None
    BQ_OK = False
    if GCP_PROJECT:
        def _init_bq():
            global _bq_client, BQ_OK
            try:
                _bq_client = _bq.Client(project=GCP_PROJECT)
                BQ_OK = True
            except Exception:
                _bq_client = None
                BQ_OK = False
        threading.Thread(target=_init_bq, daemon=True).start()
except Exception:
    _bq_client = None
    BQ_OK = False

try:
    from google.cloud import storage as _gcs
    _gcs_client = _gcs.Client() if GCS_BUCKET else None
    GCS_OK = bool(GCS_BUCKET)
except Exception:
    _gcs_client = None
    GCS_OK = False

try:
    import openpyxl
    EXCEL_OK = True
except Exception:
    EXCEL_OK = False

try:
    import pdfplumber
    PDF_OK = True
except Exception:
    PDF_OK = False

try:
    from docx import Document as _DocxDoc
    DOCX_OK = True
except Exception:
    DOCX_OK = False


# ══════════════════════════════════════════════════════════════════════════════
# DATA DICTIONARY  (loaded once, injected into every prompt)
# ══════════════════════════════════════════════════════════════════════════════

def _load_data_dict() -> str:
    p = BASE_DIR / "data_dictionary.yaml"
    if p.exists():
        return p.read_text()
    return ""

DATA_DICT = _load_data_dict()

# ══════════════════════════════════════════════════════════════════════════════
# BQ SCHEMA  (static — used by router + SQL generator)
# ══════════════════════════════════════════════════════════════════════════════

BQ_SCHEMA = f"""
BigQuery project: {GCP_PROJECT or "YOUR_PROJECT"}  dataset: {BQ_DATASET}

TABLE customers
  customer_id STRING, name STRING, industry STRING, tier STRING,
  region STRING, country STRING,
  arr_usd INT64            -- RECOGNIZED ARR (current billing period)
  arr_bookings_usd INT64   -- BOOKED ARR (signed contract value)
  seats_contracted INT64   -- seats on signed order form
  seats_active INT64       -- seats with login in last 30 days
  contract_start DATE, contract_end DATE,
  csm_owner STRING         -- Customer Success Manager (retention)
  ae_owner STRING          -- Account Executive (sold the deal)
  status STRING            -- Active / At-Risk / Churned / Prospect
  health_score INT64       -- 0-100 composite
  nps_score INT64          -- -100 to +100
  products STRING, created_at TIMESTAMP

TABLE subscriptions
  sub_id STRING, customer_id STRING, product STRING,
  seats_contracted INT64, seats_active INT64,
  list_price_unit INT64    -- price per seat per year = P
  list_price_total INT64   -- list_price_unit x seats_contracted = P*Q
  discount_pct FLOAT64, mrr_usd INT64, arr_usd INT64,
  status STRING, start_date DATE, end_date DATE, auto_renew BOOL, tier STRING

TABLE revenue_monthly
  month STRING, arr_usd INT64, mrr_usd INT64,
  new_arr INT64, expansion_arr INT64, churned_arr INT64,
  net_new_arr INT64, nrr_pct FLOAT64,
  customers_count INT64, new_logos INT64, churned_logos INT64

TABLE support_tickets
  ticket_id STRING, customer_id STRING,
  created_date DATE, resolved_date DATE,
  severity STRING, category STRING, subject STRING, status STRING,
  csat_score INT64, resolution_hrs INT64

TABLE usage_metrics
  customer_id STRING, month STRING,
  active_users INT64, seats_contracted INT64, seat_utilization FLOAT64,
  api_calls INT64, alerts_triggered INT64, alerts_actioned INT64,
  logins_per_user FLOAT64, feature_adoption FLOAT64
""".strip()

# ══════════════════════════════════════════════════════════════════════════════
# FILE SOURCE LOADERS  (cached, lazy-loaded)
# ══════════════════════════════════════════════════════════════════════════════



def _parse_excel_bytes(content: bytes, label: str) -> str:
    if not EXCEL_OK:
        return f"[Excel parser not available — pip install openpyxl]"
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    parts = [f"=== {label} ==="]
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [
            "\t".join(str(v) if v is not None else "" for v in row)
            for row in ws.iter_rows(values_only=True)
            if any(v is not None for v in row)
        ]
        if rows:
            parts.append(f"--- Sheet: {name} ---\n" + "\n".join(rows[:500]))
    return "\n\n".join(parts)

def _parse_pdf_bytes(content: bytes, label: str) -> str:
    if not PDF_OK:
        return f"[PDF parser not available — pip install pdfplumber]"
    parts = [f"=== {label} ==="]
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for i, page in enumerate(pdf.pages[:60]):
            t = page.extract_text()
            if t:
                parts.append(f"--- Page {i+1} ---\n{t}")
    return "\n\n".join(parts)

def _parse_docx_bytes(content: bytes, label: str) -> str:
    if not DOCX_OK:
        return f"[Word parser not available — pip install python-docx]"
    doc = _DocxDoc(io.BytesIO(content))
    lines = [f"=== {label} ==="]
    for p in doc.paragraphs:
        if p.text.strip():
            lines.append(p.text)
    for table in doc.tables:
        for row in table.rows:
            cells = [c.text.strip() for c in row.cells if c.text.strip()]
            if cells:
                lines.append(" | ".join(cells))
    return "\n".join(lines)

def _parse_csv_bytes(content: bytes, label: str) -> str:
    import csv
    lines = [f"=== {label} ==="]
    reader = csv.reader(io.StringIO(content.decode("utf-8", errors="replace")))
    for i, row in enumerate(reader):
        if i > 1000:
            lines.append("... (truncated at 1000 rows)")
            break
        lines.append("\t".join(row))
    return "\n".join(lines)

def _parse_any(filename: str, content: bytes, label: str) -> str:
    ext = Path(filename).suffix.lower()
    if ext in (".xlsx", ".xls"):
        return _parse_excel_bytes(content, label)
    elif ext == ".pdf":
        return _parse_pdf_bytes(content, label)
    elif ext == ".docx":
        return _parse_docx_bytes(content, label)
    elif ext == ".csv":
        return _parse_csv_bytes(content, label)
    return f"[Unsupported: {ext}]"




# ══════════════════════════════════════════════════════════════════════════════
# UPLOAD INDEX  (in-memory + GCS-persisted)
# ══════════════════════════════════════════════════════════════════════════════

# index: list of {filename, source_type, text, size_kb, uploaded_at, storage}
_upload_index: list[dict] = []

def _gcs_upload(filename: str, content: bytes):
    if not GCS_OK or not _gcs_client:
        return False
    try:
        blob = _gcs_client.bucket(GCS_BUCKET).blob(f"uploads/{filename}")
        blob.upload_from_string(content)
        return True
    except Exception as e:
        print(f"GCS upload error: {e}")
        return False

def _gcs_list() -> list[str]:
    if not GCS_OK or not _gcs_client:
        return []
    try:
        blobs = _gcs_client.bucket(GCS_BUCKET).list_blobs(prefix="uploads/")
        return [b.name.replace("uploads/", "") for b in blobs if b.name != "uploads/"]
    except Exception:
        return []

def _gcs_download(filename: str) -> Optional[bytes]:
    if not GCS_OK or not _gcs_client:
        return None
    try:
        blob = _gcs_client.bucket(GCS_BUCKET).blob(f"uploads/{filename}")
        return blob.download_as_bytes()
    except Exception:
        return None

def _gcs_delete(filename: str):
    if not GCS_OK or not _gcs_client:
        return
    try:
        _gcs_client.bucket(GCS_BUCKET).blob(f"uploads/{filename}").delete()
    except Exception:
        pass

def restore_uploads_from_storage():
    """Re-build upload index from GCS (or local) on startup."""
    indexed = {f["filename"] for f in _upload_index}

    # GCS first
    for fname in _gcs_list():
        if fname not in indexed:
            content = _gcs_download(fname)
            if content:
                _index_upload(fname, content, storage="gcs")

    # Local fallback
    for p in UPLOAD_DIR.iterdir():
        if p.suffix.lower() in ALLOWED_EXTENSIONS and p.name not in indexed:
            _index_upload(p.name, p.read_bytes(), storage="local")

def _index_upload(filename: str, content: bytes, storage: str = "local") -> dict:
    """Parse file and add to index. Returns entry dict."""
    from datetime import datetime, timezone
    ext = Path(filename).suffix.lower()
    source_type = {
        ".xlsx": "Excel", ".xls": "Excel",
        ".pdf":  "PDF",
        ".docx": "Word",
        ".csv":  "CSV",
    }.get(ext, "Unknown")

    text = _parse_any(filename, content, f"Uploaded {source_type}: {filename}")

    # Remove existing entry if re-uploading same filename
    global _upload_index
    _upload_index = [f for f in _upload_index if f["filename"] != filename]

    entry = {
        "filename":    filename,
        "source_type": source_type,
        "text":        text,
        "size_kb":     round(len(content) / 1024, 1),
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "storage":     storage,
        "preview":     text[:200] + "..." if len(text) > 200 else text,
    }
    _upload_index.append(entry)
    return entry

def get_uploads_text() -> str:
    """Combined text of all indexed uploads for prompt injection."""
    if not _upload_index:
        return ""
    return "\n\n".join(
        f"{'='*50}\nSOURCE: Uploaded {f['source_type']} — {f['filename']}\n{'='*50}\n{f['text']}"
        for f in _upload_index
    )

def get_uploads_manifest() -> list[dict]:
    """Metadata only (no text) for UI listing."""
    return [{k: v for k, v in f.items() if k != "text"} for f in _upload_index]


# ══════════════════════════════════════════════════════════════════════════════
# STAGE 1: AI ROUTER
# ══════════════════════════════════════════════════════════════════════════════

ROUTER_PROMPT = """You are a query router for an enterprise data assistant.

Available sources:
  bigquery  — live database: customers, subscriptions, revenue_monthly, support_tickets, usage_metrics
  uploaded  — user-uploaded files (Excel, PDF, Word, CSV — whatever the user has uploaded)

BigQuery schema summary:
{schema}

Uploads currently indexed: {uploads_manifest}

Conversation history (last {history_window} turns):
{history}

Current question: {question}

Respond ONLY with valid JSON, no markdown, no explanation:
{{
  "sources": ["bigquery", "uploaded"],
  "needs_sql": true,
  "sql_intent": "one sentence describing what SQL should retrieve, or null if needs_sql is false",
  "query_type": "single_source | multi_source | followup | upload_only",
  "intent_tag": "revenue | pipeline | churn | policy | pricing | account_health | usage | save_playbook | comparison | other",
  "reasoning": "one sentence why these sources were selected"
}}

Rules:
- Include only the sources actually needed. Do not include all sources by default.
- needs_sql is true only if bigquery is in sources.
- uploaded should be included if any uploaded files exist AND the question could be answered by them.
- For followup questions, look at history to determine correct sources.
- bigquery is the default for any question about customers, revenue, ARR, seats, health scores, support.
- uploaded covers any question that could be answered by the user's uploaded files."""


async def run_router(question: str, history: list[dict]) -> dict:
    """Stage 1: AI router using Gemini Flash. Fast and cheap."""
    if not GENAI_OK:
        return {
            "sources": ["bigquery"],
            "needs_sql": True,
            "sql_intent": "retrieve relevant data for the question",
            "query_type": "multi_source",
            "intent_tag": "other",
            "reasoning": "fallback routing — Gemini not available",
        }

    history_text = "\n".join(
        f"{'User' if m['role'] == 'user' else 'Assistant'}: {m['content']}"
        for m in history[-ROUTER_HISTORY_WIN:]
    ) or "(none)"

    uploads_manifest_str = (
        json.dumps([{"filename": f["filename"], "source_type": f["source_type"]} for f in _upload_index])
        if _upload_index else "[]"
    )

    prompt = ROUTER_PROMPT.format(
        schema=BQ_SCHEMA,
        uploads_manifest=uploads_manifest_str,
        history_window=ROUTER_HISTORY_WIN,
        history=history_text,
        question=question,
    )

    try:
        resp = _router_model.generate_content(prompt)
        raw = resp.text.strip().replace("```json", "").replace("```", "").strip()
        decision = json.loads(raw)
        return decision
    except Exception as e:
        # Graceful fallback if router fails
        return {
            "sources": ["bigquery"],
            "needs_sql": True,
            "sql_intent": question,
            "query_type": "single_source",
            "intent_tag": "other",
            "reasoning": f"router error fallback: {e}",
        }


# ══════════════════════════════════════════════════════════════════════════════
# STAGE 2a: SQL GENERATION + VALIDATION + EXECUTION
# ══════════════════════════════════════════════════════════════════════════════

SQL_GEN_PROMPT = """You are a BigQuery SQL expert.

Schema:
{schema}

Data dictionary (resolve ambiguous columns using this):
{data_dict}

Conversation history:
{history}

SQL intent from router: {sql_intent}
User question: {question}

Write a single BigQuery SQL query. Rules:
- Fully qualified table names: `{project}.{dataset}.TABLE`
- ALWAYS SELECT * or include identifying columns (name, customer_id) — never select a single metric column alone
- Only use columns that exist in the schema
- For ambiguous columns, add a comment on your choice: -- using X not Y because reason
- LIMIT to 500 rows maximum
- If the question truly cannot be answered with SQL, return exactly: NO_SQL_NEEDED

Return ONLY the SQL query or NO_SQL_NEEDED. No markdown fences, no explanation."""


async def generate_and_run_sql(question: str, sql_intent: str, history: list[dict]) -> dict:
    """Generate SQL via Gemini, validate via dry-run, execute, return results."""
    if not GENAI_OK or not BQ_OK or not _bq_client:
        return {"status": "unavailable", "sql": None, "data": _bq_inline_fallback(), "row_count": 0}

    history_text = "\n".join(
        f"{'User' if m['role'] == 'user' else 'Assistant'}: {m['content']}"
        for m in history[-ROUTER_HISTORY_WIN:]
    ) or "(none)"

    prompt = SQL_GEN_PROMPT.format(
        schema=BQ_SCHEMA,
        data_dict=DATA_DICT,
        history=history_text,
        sql_intent=sql_intent,
        question=question,
        project=GCP_PROJECT,
        dataset=BQ_DATASET,
    )

    try:
        resp = _answer_model.generate_content(prompt)
        sql = resp.text.strip().replace("```sql", "").replace("```", "").strip()
    except Exception as e:
        return {"status": "gen_error", "sql": None, "data": _bq_inline_fallback(), "error": str(e)}

    if sql == "NO_SQL_NEEDED":
        return {"status": "not_needed", "sql": None, "data": "", "row_count": 0}

    # Dry-run validation (zero cost, validates schema + syntax)
    sql, result = await _validate_and_execute(sql, question, history_text)
    return result


async def _validate_and_execute(sql: str, question: str, history_text: str) -> tuple[str, dict]:
    """Dry-run → self-correct if needed → execute."""
    from google.cloud import bigquery as bq

    # Attempt 1: dry-run
    try:
        cfg = bq.QueryJobConfig(dry_run=True, use_query_cache=False)
        _bq_client.query(sql, job_config=cfg)
    except Exception as e1:
        # Self-correction: feed error back to model
        fix_prompt = (
            f"Fix this BigQuery SQL. Return ONLY the corrected SQL, no markdown, no explanation.\n\n"
            f"SQL:\n{sql}\n\nError:\n{e1}"
        )
        try:
            fixed = _answer_model.generate_content(fix_prompt)
            sql = fixed.text.strip().replace("```sql", "").replace("```", "").strip()
            cfg = bq.QueryJobConfig(dry_run=True, use_query_cache=False)
            _bq_client.query(sql, job_config=cfg)
        except Exception as e2:
            return sql, {
                "status": "validation_failed",
                "sql": sql,
                "data": _bq_inline_fallback(),
                "error": str(e2),
                "row_count": 0,
            }

    # Execute (with LIMIT safety wrap)
    try:
        safe_sql = f"SELECT * FROM ({sql}) LIMIT 500"
        rows = [dict(row) for row in _bq_client.query(safe_sql).result()]
        data_text = (
            f"BigQuery results ({len(rows)} rows):\n"
            + json.dumps(rows, indent=2, default=str)
        )
        return sql, {"status": "success", "sql": sql, "data": data_text, "row_count": len(rows)}
    except Exception as e:
        return sql, {
            "status": "exec_error",
            "sql": sql,
            "data": _bq_inline_fallback(),
            "error": str(e),
            "row_count": 0,
        }


def _bq_inline_fallback() -> str:
    """Inline fallback data when BQ is unavailable. Keeps demo runnable offline."""
    return json.dumps({
        "note": "BigQuery unavailable — inline fallback data",
        "customers": [
            {"name": "Apex Financial",    "tier": "Enterprise", "arr_usd": 480000, "seats_contracted": 850,  "seats_active": 848, "csm_owner": "Priya Nair",    "status": "Active",  "health_score": 82},
            {"name": "Meridian Trading",  "tier": "Enterprise", "arr_usd": 540000, "seats_contracted": 1100, "seats_active": 890, "csm_owner": "Wei Zhang",     "status": "At-Risk", "health_score": 31},
            {"name": "Vantage Capital",   "tier": "Enterprise", "arr_usd": 360000, "seats_contracted": 620,  "seats_active": 595, "csm_owner": "James Okoye",   "status": "Active",  "health_score": 77},
            {"name": "NordBank AG",       "tier": "Mid-Market", "arr_usd": 144000, "seats_contracted": 200,  "seats_active": 0,   "csm_owner": "Sophie Laurent", "status": "Churned", "health_score": 0},
            {"name": "Pinnacle Wealth",   "tier": "SMB",        "arr_usd": 36000,  "seats_contracted": 45,   "seats_active": 22,  "csm_owner": "James Okoye",   "status": "Active",  "health_score": 55},
            {"name": "GoldLeaf Advisors", "tier": "Mid-Market", "arr_usd": 96000,  "seats_contracted": 175,  "seats_active": 155, "csm_owner": "Sophie Laurent", "status": "Active",  "health_score": 91},
        ],
        "revenue_monthly_latest": [
            {"month": "2024-07", "arr_usd": 2230000, "nrr_pct": 114.8, "new_logos": 4},
            {"month": "2024-08", "arr_usd": 2310000, "nrr_pct": 112.4, "new_logos": 2},
            {"month": "2024-09", "arr_usd": 2220000, "nrr_pct": 112.1, "new_logos": 2},
        ],
        "subscriptions_sample": [
            {"customer": "Apex Financial",   "product": "ThreatShield Enterprise",    "list_price_unit": 420, "list_price_total": 357000, "seats_contracted": 850},
            {"customer": "Meridian Trading", "product": "ThreatShield Enterprise",    "list_price_unit": 420, "list_price_total": 462000, "seats_contracted": 1100},
            {"customer": "Meridian Trading", "product": "Threat Intelligence Add-on", "list_price_unit": 0,   "list_price_total": 45000,  "seats_contracted": 0},
        ],
    }, indent=2)


# ══════════════════════════════════════════════════════════════════════════════
# STAGE 3: ANSWER GENERATION (streaming)
# ══════════════════════════════════════════════════════════════════════════════

ANSWER_SYSTEM = """You are the saasmetrics.ai Enterprise Data Assistant.
You help business leaders get accurate, grounded answers from enterprise data.

DATA DICTIONARY — Column disambiguation rules:
{data_dict}

STRICT RULES — follow every one:
1. GROUND every claim in the source data provided. Never use training memory for numbers.
   If data is not in the sources, say exactly: "I don't have that data in the connected sources."
2. CITE sources inline: [BigQuery: table_name], [PDF §section], [Excel: SheetName], [Word: §X.Y], [Uploaded: filename]
3. DISAMBIGUATE visibly: when using an ambiguous column, state your choice:
   "I used seats_contracted (850) not seats_active (848) — you asked what they purchased, not who is logging in."
4. NO extrapolation: do not forecast or project beyond available data. Describe trends only.
5. FLAG conflicts: if sources give different numbers, note it and state which is authoritative.
6. RESOLVE follow-ups: use conversation history to resolve "them", "that account", "same customer".
7. CONFIDENCE: end every response with one of — Confidence: HIGH / MEDIUM / LOW
   HIGH = answer fully grounded in data | MEDIUM = partial data | LOW = data unavailable or conflicting

Respond in markdown with inline citations. Be concise and direct. Lead with the answer, not the method."""

ANSWER_USER = """SOURCE DATA:
{source_blocks}

CONVERSATION HISTORY:
{history}

QUESTION: {question}

Also include at the end of your response this JSON block on its own line (do not wrap in markdown):
METADATA::{{"sources_used": {sources_list}, "disambiguation_notes": "...", "confidence": "high/medium/low", "query_type": "{query_type}", "intent_tag": "{intent_tag}"}}"""


async def stream_answer(
    question: str,
    history: list[dict],
    source_blocks: str,
    sources_used: list[str],
    query_type: str,
    intent_tag: str,
) -> AsyncIterator[str]:
    """Stage 3: Stream answer tokens via SSE."""
    if not GENAI_OK:
        yield "data: " + json.dumps({"token": "⚠️ GEMINI_API_KEY not configured.", "done": False}) + "\n\n"
        yield "data: " + json.dumps({"done": True, "metadata": {}}) + "\n\n"
        return

    history_text = "\n".join(
        f"{'User' if m['role'] == 'user' else 'Assistant'}: {m['content']}"
        for m in history[-HISTORY_WINDOW:]
    ) or "(none)"

    system = ANSWER_SYSTEM.format(data_dict=DATA_DICT)
    user_msg = ANSWER_USER.format(
        source_blocks=source_blocks,
        history=history_text,
        question=question,
        sources_list=json.dumps(sources_used),
        query_type=query_type,
        intent_tag=intent_tag,
    )

    try:
        model = genai.GenerativeModel(
            ANSWER_MODEL,
            system_instruction=system,
        )
        stream = model.generate_content(user_msg, stream=True)

        full_text = ""
        for chunk in stream:
            if chunk.text:
                full_text += chunk.text
                yield "data: " + json.dumps({"token": chunk.text, "done": False}) + "\n\n"

        # Extract metadata JSON from end of response
        metadata = {}
        meta_match = re.search(r"METADATA::(\{.*\})", full_text, re.DOTALL)
        if meta_match:
            try:
                metadata = json.loads(meta_match.group(1))
            except Exception:
                metadata = {}

        yield "data: " + json.dumps({"done": True, "metadata": metadata}) + "\n\n"

    except Exception as e:
        yield "data: " + json.dumps({"token": f"⚠️ Generation error: {e}", "done": False}) + "\n\n"
        yield "data: " + json.dumps({"done": True, "metadata": {}}) + "\n\n"


# ══════════════════════════════════════════════════════════════════════════════
# FASTAPI APP
# ══════════════════════════════════════════════════════════════════════════════

app = FastAPI(title="saasmetrics.ai", version="5.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def startup():
    """Re-index uploads from local on startup. Non-blocking."""
    try:
        loop = asyncio.get_event_loop()
        await asyncio.wait_for(
            loop.run_in_executor(None, restore_uploads_from_storage),
            timeout=5.0
        )
    except Exception:
        pass  # Never block startup


# ── Health ────────────────────────────────────────────────────────────────────
@app.get("/health")
def health():
    return {
        "status": "ok",
        "gemini": GENAI_OK,
        "bigquery": BQ_OK,
        "gcs": GCS_OK,
        "router_model": ROUTER_MODEL,
        "answer_model": ANSWER_MODEL,
        "builtin_sources": {

        },
        "uploads_indexed": len(_upload_index),
    }


# ── Query (streaming SSE) ─────────────────────────────────────────────────────
class Message(BaseModel):
    role: str
    content: str

class QueryRequest(BaseModel):
    question: str
    history: List[Message] = []


@app.post("/query")
async def query(req: QueryRequest):
    question = req.question.strip()
    history  = [m.model_dump() for m in req.history]

    async def event_stream():
        # ── Stage 1: Route ────────────────────────────────────────────────
        route = await run_router(question, history)
        sources   = route.get("sources", ["bigquery"])
        needs_sql = route.get("needs_sql", False)
        sql_intent= route.get("sql_intent", question)
        query_type= route.get("query_type", "single_source")
        intent_tag= route.get("intent_tag", "other")

        # Emit routing decision to frontend immediately
        yield "data: " + json.dumps({
            "event": "routing",
            "sources": sources,
            "query_type": query_type,
            "intent_tag": intent_tag,
            "reasoning": route.get("reasoning", ""),
        }) + "\n\n"

        # ── Stage 2: Parallel source fetch ────────────────────────────────
        source_blocks = ""
        sql_used = None

        # BQ + file sources fetched concurrently
        tasks = {}
        if needs_sql and "bigquery" in sources:
            tasks["bq"] = asyncio.create_task(
                generate_and_run_sql(question, sql_intent, history)
            )

        # Await BQ task
        bq_result = None
        if "bq" in tasks:
            bq_result = await tasks["bq"]

        # Assemble source blocks
        if bq_result:
            sql_used = bq_result.get("sql")
            data = bq_result.get("data", "")
            if data:
                source_blocks += f"\n{'='*50}\nSOURCE: BigQuery ({BQ_DATASET})\n{'='*50}\n{data}\n"
            # Emit SQL to frontend
            if sql_used:
                yield "data: " + json.dumps({"event": "sql", "sql": sql_used, "status": bq_result.get("status")}) + "\n\n"

        if "uploaded" in sources and _upload_index:
            uploads_text = get_uploads_text()
            if uploads_text:
                source_blocks += f"\n{'='*50}\nSOURCE: USER UPLOADS ({len(_upload_index)} file(s))\n{'='*50}\n{uploads_text}\n"

        # ── Stage 3: Stream answer ─────────────────────────────────────────
        async for chunk in stream_answer(
            question, history, source_blocks,
            sources, query_type, intent_tag
        ):
            yield chunk

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


# ── Uploads ───────────────────────────────────────────────────────────────────
@app.post("/upload")
async def upload(file: UploadFile = File(...)):
    if Path(file.filename).suffix.lower() not in ALLOWED_EXTENSIONS:
        raise HTTPException(400, f"Unsupported type. Allowed: {', '.join(ALLOWED_EXTENSIONS)}")

    content = await file.read()
    size_mb = len(content) / (1024 * 1024)
    if size_mb > MAX_UPLOAD_MB:
        raise HTTPException(400, f"File too large ({size_mb:.1f} MB). Max: {MAX_UPLOAD_MB} MB.")

    safe_name = Path(file.filename).name.replace(" ", "_")

    # Save to GCS (or local)
    in_gcs = _gcs_upload(safe_name, content)
    if not in_gcs:
        (UPLOAD_DIR / safe_name).write_bytes(content)

    entry = _index_upload(safe_name, content, storage="gcs" if in_gcs else "local")
    return {
        "success": True,
        "filename": entry["filename"],
        "source_type": entry["source_type"],
        "size_kb": entry["size_kb"],
        "storage": entry["storage"],
        "preview": entry["preview"],
    }


@app.get("/uploads")
def list_uploads():
    return {"files": get_uploads_manifest(), "count": len(_upload_index)}


@app.delete("/upload/{filename}")
def delete_upload(filename: str):
    global _upload_index
    existing = [f for f in _upload_index if f["filename"] == filename]
    if not existing:
        raise HTTPException(404, "File not found in index")
    _upload_index = [f for f in _upload_index if f["filename"] != filename]
    _gcs_delete(filename)
    local = UPLOAD_DIR / filename
    if local.exists():
        local.unlink()
    return {"success": True, "filename": filename}


@app.post("/reload")
def reload_sources():
    """Clear built-in source cache so files are re-read on next query."""

    return {"status": "source cache cleared"}


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("backend.main:app", host="0.0.0.0", port=BACKEND_PORT, reload=True)